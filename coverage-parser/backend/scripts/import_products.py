"""
保险产品库批量导入脚本
从Excel文件导入产品数据到数据库
"""

import openpyxl
import sys
import os
from datetime import datetime

# 添加项目根目录到Python路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def read_excel(file_path):
    """读取Excel文件"""
    print(f"📂 正在读取Excel文件: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"✅ 文件读取成功！")
        print(f"   工作表名称: {ws.title}")
        print(f"   总行数: {ws.max_row}")
        print(f"   总列数: {ws.max_column}")
        
        # 读取表头
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col).value
            headers.append(cell_value)
        
        print(f"\n📋 列名: {headers}")
        
        # 读取数据
        products = []
        for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
            row_data = {}
            for col, header in enumerate(headers, 1):
                cell_value = ws.cell(row, col).value
                row_data[header] = cell_value
            products.append(row_data)
            
            # 每1000行打印进度
            if row % 1000 == 0:
                print(f"   已读取 {row-1} 行...")
        
        print(f"\n✅ 成功读取 {len(products)} 条产品数据！")
        return products, headers
        
    except Exception as e:
        print(f"❌ 读取Excel文件失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def map_product_category(category):
    """映射产品大类到policyType"""
    mapping = {
        "疾病险": "重疾险",
        "人寿险": "人寿险",
        "意外险": "意外险",
        "年金险": "年金险"
    }
    return mapping.get(category, category)


def generate_sql(products, filter_category=None):
    """生成SQL插入语句
    
    Args:
        products: 产品列表
        filter_category: 筛选的产品大类（如：'疾病险'），None表示导入全部
    """
    print(f"\n📝 正在生成SQL语句...")
    if filter_category:
        print(f"   筛选条件：仅导入 {filter_category}")
    
    sql_statements = []
    
    # 清空表（可选）
    # sql_statements.append("DELETE FROM insurance_product_library;")
    
    filtered_count = 0
    for i, product in enumerate(products, 1):
        policyId = product.get('保险产品ID号', '').strip()
        company = product.get('公司名称', '').strip()
        productName = product.get('保险产品名称', '').strip()
        category = product.get('保险大类', '').strip()
        subCategory = product.get('保险小类', '').strip() if product.get('保险小类') else None
        coveragePeriod = product.get('保障期限', '').strip() if product.get('保障期限') else None
        paymentPeriod = product.get('交费期限', '').strip() if product.get('交费期限') else None
        salesStatus = product.get('销售状态', '在售').strip()
        
        # 跳过空行
        if not policyId or not company or not productName:
            continue
        
        # 筛选产品大类
        if filter_category and category != filter_category:
            filtered_count += 1
            continue
        
        # 转义单引号
        policyId = policyId.replace("'", "''")
        company = company.replace("'", "''")
        productName = productName.replace("'", "''")
        category = category.replace("'", "''")
        if subCategory:
            subCategory = subCategory.replace("'", "''")
        if coveragePeriod:
            coveragePeriod = coveragePeriod.replace("'", "''")
        if paymentPeriod:
            paymentPeriod = paymentPeriod.replace("'", "''")
        salesStatus = salesStatus.replace("'", "''")
        
        # 映射policyType
        policyType = map_product_category(category)
        
        sql = f"""
INSERT INTO insurance_product_library 
  (policyId, insuranceCompany, productName, productCategory, productSubCategory, 
   coveragePeriod, paymentPeriod, salesStatus, policyType, createdAt, updatedAt)
VALUES 
  ('{policyId}', '{company}', '{productName}', '{category}', {f"'{subCategory}'" if subCategory else 'NULL'}, 
   {f"'{coveragePeriod}'" if coveragePeriod else 'NULL'}, {f"'{paymentPeriod}'" if paymentPeriod else 'NULL'}, 
   '{salesStatus}', '{policyType}', NOW(), NOW())
ON CONFLICT (policyId) DO UPDATE SET
  insuranceCompany = EXCLUDED.insuranceCompany,
  productName = EXCLUDED.productName,
  productCategory = EXCLUDED.productCategory,
  productSubCategory = EXCLUDED.productSubCategory,
  coveragePeriod = EXCLUDED.coveragePeriod,
  paymentPeriod = EXCLUDED.paymentPeriod,
  salesStatus = EXCLUDED.salesStatus,
  policyType = EXCLUDED.policyType,
  updatedAt = NOW();
"""
        sql_statements.append(sql.strip())
        
        # 每1000条打印进度
        if len(sql_statements) % 1000 == 0:
            print(f"   已生成 {len(sql_statements)} 条SQL...")
    
    if filter_category:
        print(f"   已过滤 {filtered_count} 条非{filter_category}产品")
    print(f"\n✅ 成功生成 {len(sql_statements)} 条SQL语句！")
    return sql_statements


def save_sql_file(sql_statements, output_file):
    """保存SQL到文件"""
    print(f"\n💾 正在保存SQL文件: {output_file}")
    
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("-- 保险产品库导入SQL\n")
            f.write(f"-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"-- 总记录数: {len(sql_statements)}\n\n")
            f.write("BEGIN;\n\n")
            
            for sql in sql_statements:
                f.write(sql + "\n\n")
            
            f.write("COMMIT;\n")
        
        print(f"✅ SQL文件保存成功！")
        print(f"   文件路径: {output_file}")
        print(f"   文件大小: {os.path.getsize(output_file) / 1024:.2f} KB")
        
    except Exception as e:
        print(f"❌ 保存SQL文件失败: {e}")
        sys.exit(1)


def main():
    """主函数"""
    print("=" * 60)
    print("📦 保险产品库批量导入脚本")
    print("=" * 60)
    
    # Excel文件路径
    excel_file = "/Users/hanyang/Desktop/保险解析助手/保险库.xlsx"
    output_sql = "/Users/hanyang/Desktop/保险解析助手/coverage-parser/backend/scripts/import_products.sql"
    
    # ⚠️ 筛选条件：只导入疾病险
    filter_category = "疾病险"  # 可选：None（全部）、"疾病险"、"人寿险"、"意外险"、"年金险"
    
    if filter_category:
        print(f"\n⚠️  筛选模式：仅导入 [{filter_category}] 类产品")
        output_sql = output_sql.replace('.sql', f'_{filter_category}.sql')
    
    # 1. 读取Excel
    products, headers = read_excel(excel_file)
    
    # 2. 生成SQL
    sql_statements = generate_sql(products, filter_category)
    
    # 3. 保存SQL文件
    save_sql_file(sql_statements, output_sql)
    
    print("\n" + "=" * 60)
    print("✅ 导入脚本执行完成！")
    print("=" * 60)
    print(f"\n📝 下一步操作：")
    print(f"   1. 检查生成的SQL文件: {output_sql}")
    print(f"   2. 执行SQL导入数据库")
    print(f"   3. 验证数据是否正确导入")
    print()


if __name__ == "__main__":
    main()

