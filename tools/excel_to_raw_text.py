#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从Excel提取原文数据，生成标准的 原文条款-批次X.md

功能：
  - 读取Excel指定工作表
  - 提取序号、保单ID号、责任类型、责任名称、责任原文
  - 生成标准格式的原文条款文件

使用方法：
  python excel_to_raw_text.py --input 责任库导出-模版.xlsx --sheet 疾病责任 --batch 16
  python excel_to_raw_text.py --input 责任库导出-模版.xlsx --sheet 疾病责任 --output 原文条款-批次16.md
"""

import argparse
from pathlib import Path
from openpyxl import load_workbook


def extract_raw_text(input_file: str, sheet_name: str, output_file: str = None, batch_num: int = None):
    """
    从Excel提取原文数据
    
    Args:
        input_file: Excel文件路径
        sheet_name: 工作表名称
        output_file: 输出文件路径（可选）
        batch_num: 批次编号（可选，用于自动生成输出文件名）
    """
    
    # 加载Excel
    print("="*80)
    print("从Excel提取原文数据")
    print("="*80)
    print(f"输入文件: {input_file}")
    print(f"工作表: {sheet_name}")
    print()
    
    wb = load_workbook(input_file, read_only=True)
    
    # 检查工作表是否存在
    if sheet_name not in wb.sheetnames:
        print(f"❌ 错误: 工作表 '{sheet_name}' 不存在")
        print(f"可用的工作表: {', '.join(wb.sheetnames)}")
        wb.close()
        return
    
    ws = wb[sheet_name]
    
    # 读取表头（第1行）
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print(f"表头: {headers}")
    print()
    
    # 查找关键列的索引
    col_indices = {}
    required_cols = ['序号', '保单ID号', '责任类型', '责任名称', '责任原文']
    
    for i, header in enumerate(headers):
        if header in required_cols:
            col_indices[header] = i
    
    # 检查是否所有必需列都存在
    missing_cols = [col for col in required_cols if col not in col_indices]
    if missing_cols:
        print(f"❌ 错误: 缺少必需列: {', '.join(missing_cols)}")
        wb.close()
        return
    
    print(f"✅ 找到所有必需列")
    for col, idx in col_indices.items():
        print(f"  {col}: 第{idx+1}列")
    print()
    
    # 提取数据
    print("提取数据...")
    data_rows = []
    序号_list = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        序号 = row[col_indices['序号']]
        
        # 跳过空行
        if not 序号:
            continue
        
        保单ID号 = row[col_indices['保单ID号']] or ''
        责任类型 = row[col_indices['责任类型']] or ''
        责任名称 = row[col_indices['责任名称']] or ''
        责任原文 = row[col_indices['责任原文']] or ''
        
        # 转换序号为整数
        try:
            序号 = int(序号)
        except:
            print(f"  ⚠️ 第{row_idx}行: 序号格式错误，跳过")
            continue
        
        序号_list.append(序号)
        
        data_rows.append({
            '序号': 序号,
            '保单ID号': 保单ID号,
            '责任类型': 责任类型,
            '责任名称': 责任名称,
            '责任原文': 责任原文
        })
    
    wb.close()
    
    if not data_rows:
        print("❌ 错误: 没有提取到任何数据")
        return
    
    print(f"✅ 提取到 {len(data_rows)} 条数据")
    print(f"   序号范围: {min(序号_list)} - {max(序号_list)}")
    print()
    
    # 确定输出文件路径
    if not output_file:
        if batch_num:
            output_file = f"原文条款/原文条款-批次{batch_num}.md"
        else:
            # 根据序号范围自动推断批次
            min_seq = min(序号_list)
            max_seq = max(序号_list)
            inferred_batch = (min_seq - 1) // 200 + 1
            output_file = f"原文条款/原文条款-批次{inferred_batch}.md"
            print(f"⚠️ 未指定批次编号，根据序号范围推断为批次{inferred_batch}")
    
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 生成原文条款文件
    print(f"生成原文条款文件: {output_file}")
    
    with open(output_path, 'w', encoding='utf-8') as f:
        # 写入文件头
        min_seq = min(序号_list)
        max_seq = max(序号_list)
        f.write(f"# 原文条款 - 批次{batch_num or (min_seq-1)//200+1}\n\n")
        f.write(f"序号范围: {min_seq}-{max_seq}\n")
        f.write(f"数据来源: {Path(input_file).name} - {sheet_name}\n")
        f.write(f"总数: {len(data_rows)} 条\n\n")
        f.write("---\n\n")
        
        # 写入数据（使用 ||| 分隔符，与历史格式一致）
        for row in data_rows:
            line = f"{row['序号']}|||{row['保单ID号']}|||{row['责任类型']}|||{row['责任名称']}|||{row['责任原文']}\n"
            f.write(line)
    
    print(f"✅ 已保存到: {output_path.absolute()}")
    print()
    print("="*80)
    print("✅ 提取完成！")
    print("="*80)
    print()
    print("下一步:")
    print(f"  cd tools")
    print(f"  python 1_generate_batch.py --batch {batch_num or (min_seq-1)//200+1}")


def main():
    parser = argparse.ArgumentParser(description='从Excel提取原文数据')
    parser.add_argument('--input', required=True, help='Excel文件路径')
    parser.add_argument('--sheet', required=True, help='工作表名称')
    parser.add_argument('--output', help='输出文件路径（可选）')
    parser.add_argument('--batch', type=int, help='批次编号（可选）')
    
    args = parser.parse_args()
    
    extract_raw_text(
        input_file=args.input,
        sheet_name=args.sheet,
        output_file=args.output,
        batch_num=args.batch
    )


if __name__ == '__main__':
    main()
